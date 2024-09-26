import os
from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from werkzeug.utils import secure_filename
import pytesseract
from PIL import Image
import cv2
import openpyxl
import json
from PIL import Image, ImageEnhance, ImageFilter
import openai
from google.cloud import vision
import io
import re
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
from openpyxl import Workbook
import os
from dotenv import load_dotenv
import subprocess
import platform
import PyPDF2

load_dotenv()  

# openai.api_key = os.getenv('OPENAI_API_KEY')

open.ai_key = 'API KEY HERE'

# Flask app configuration
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['PROCESSED_FOLDER'] = 'processed/'
app.config['OUTPUT_FOLDER'] = 'output/'
app.config['SECRET_KEY'] = 'your_secret_key'

 

# Ensure folders exist
for folder in [app.config['UPLOAD_FOLDER'], app.config['PROCESSED_FOLDER'], app.config['OUTPUT_FOLDER']]:
    os.makedirs(folder, exist_ok=True)

def split_pdf_into_pages(pdf_path):
    """Split a PDF file into single-page PDFs and return the list of file paths."""
    pdf_reader = PyPDF2.PdfReader(pdf_path)
    pdf_pages = []

    for page_num in range(len(pdf_reader.pages)):
        pdf_writer = PyPDF2.PdfWriter()
        pdf_writer.add_page(pdf_reader.pages[page_num])

        # Save each page as a separate PDF file
        page_filename = f"{os.path.splitext(pdf_path)[0]}_page_{page_num + 1}.pdf"
        with open(page_filename, 'wb') as output_pdf:
            pdf_writer.write(output_pdf)
        pdf_pages.append(page_filename)

    return pdf_pages

def merge_invoice_data(invoices):
    """Merge invoice items for pages belonging to the same supplier (case-insensitive)."""
    merged_invoices = {}
    
    for invoice in invoices:
        supplier_name = invoice.get("Supplier Name", "Unknown Supplier").strip().lower()  # Normalize case and strip spaces
        
        if supplier_name not in merged_invoices:
            merged_invoices[supplier_name] = {
                "Supplier Name": invoice.get("Supplier Name", "Unknown Supplier"),  # Use original case for output
                "Invoice Number": invoice.get("Invoice Number", "Unknown Invoice"),
                "Date": invoice.get("Date", "Unknown Date"),
                "items": invoice.get("items", [])
            }
        else:
            # Merge items if the supplier is the same (case-insensitive match)
            merged_invoices[supplier_name]["items"].extend(invoice.get("items", []))
    
    return list(merged_invoices.values())



def open_excel_file(file_path):
    try:
        if platform.system() == "Windows":
            os.startfile(file_path)
        elif platform.system() == "Darwin":  # macOS
            subprocess.run(["open", file_path])
        else:  # Assume Linux
            subprocess.run(["xdg-open", file_path])
    except Exception as e:
        print(f"Error opening file: {e}")

def convert_to_jpg(image_path):
    """Convert image or PDF to JPG format if it's not already in that format."""
    file_extension = os.path.splitext(image_path)[1].lower()
    
    if file_extension == '.pdf':
        from pdf2image import convert_from_path
        images = convert_from_path(image_path)
        jpg_image_path = os.path.splitext(image_path)[0] + '.jpg'
        images[0].save(jpg_image_path, 'JPEG')
        return jpg_image_path
    elif file_extension in ['.png', '.jpeg', '.jpg']:
        img = Image.open(image_path)
        jpg_image_path = os.path.splitext(image_path)[0] + '.jpg'
        img.convert('RGB').save(jpg_image_path, 'JPEG')
        return jpg_image_path
    else:
        raise ValueError("Unsupported file type for conversion.")

def preprocess_image(image_path):
    """Preprocess the image to enhance OCR accuracy."""
    # Load the image
    image = cv2.imread(image_path)
    # Convert to grayscale
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    # Apply Gaussian blur to reduce noise
    blurred_image = cv2.GaussianBlur(gray_image, (5, 5), 0)
    # Apply adaptive thresholding
    _, binary_image = cv2.threshold(blurred_image, 150, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    return binary_image

def perform_ocr(image):
    """Perform OCR on the preprocessed image."""
    # Use Tesseract to extract text from the image
    custom_config = r'--psm 6 --oem 3 -c preserve_interword_spaces=1'
    extracted_text = pytesseract.image_to_string(image, config=custom_config)
    return extracted_text

def parse_invoice_with_genai(extracted_text):
    """Use OpenAI to extract invoice details in JSON format."""
    prompt = (
        "Follow the JSON format below for the output. This format MUST be strictly followed:\n"
        "{\n"
        "  \"Supplier Name\": \"Supplier Name\",\n"
        "  \"Invoice Number\": \"Invoice Number\",\n"
        "  \"Date\": \"Invoice Date\",\n"
        "  \"items\": [\n"
        "    {\n"
        "      \"Description\": \"Item Name\",\n"
        "      \"Quantity\": \"Quantity\",\n"
        "      \"Rate\": \"Rate\",\n"
        "      \"GST\": \"GST\",\n"
        "      \"Sales Amount\": \"Sales Amount\"\n"
        "    }\n"
        "  ]\n"
        "}\n"
        "Extract the following information from the text:\n"
        "- Supplier Name: This should be a standalone company name and should not be confused with sender or receiver names.\n"
        "- Product Name: must be fetched completely.\n"
        "- Invoice Number: This could be labeled as Bill No., Invoice No., etc. (NOT the tax number, GSTIN/UIN, etc.).\n"
        "- Quantity: include the unit of the quantity (would NEVER EVER exist in the product description) This could be labeled as Bill Qty, Qty, EA, Total Pcs. , Tot Pcs,(str, nos, box) etc., If quantity is written in a hyphenated format like 0-1 STR, consider the second number (1). Quantity can also be identified as EA. Best to calculate the quantity by dividing the sales amount by the rate.\n"
        "- Rate: Could also be labelled as S.Rate (but never as Gross.Amt, and never the Taxable value).\n"
        "- GST: This could be labeled as GST rate or otherwise calculated by adding CSST rate + SGST percentage rate. We need the rate as a whole number, not the amount. It will be the same throughout the invoice for all items, just the sum of both (not each of them twice).\n"
        "- Sales Amount: This could be labeled as Net amount or Inv Amt, Net.Amt etc. Not to be calculateed, but rather to be picked from the data. This value is usually the last or the second last value of the data per item. (this is never the N.Rate\n\n"
        "Please note:\n"
        "- These are the only needed headers; do not include anything else.\n"
        "- Do not generate any commas, or anything that will cause error when parsing the json decoding\n"
        "- If there are multiple items in the invoice, provide an array of objects, each representing a single item with its details.\n"
        "- The Invoice Number must be correctly extracted even if it is in a tabular format or not the most prominent number. It should not be confused with other numbers in tables or addresses.\n"
        "- For entries with null values, ignore them in the output.\n"
        # "  IMPORTANT: Remeber that the quantity you extract will be a standalone entity, and wont be ever extracted from the product name. eg: if 'Exo Bar 50g(+10g) Fr Ging NB (216 pc)' is the product name, '216' will never be the quantity.\n"
        "- Lastly, since the raw text is OCR fetched, there sometimes might be issues in fetching the quantity, therefore you must calculate the quantity in this case by dividing the sales amount (WITHOUT TAX) with the rate.\n"
        "- Provide a RAW JSON OUTPUT, with no backticks or formatting (also ensure that there are no json curly braces anywhere in the parsed data as it may disturb the json structure). The supplier name and invoice number should be the headers, and the items should contain the rest.\n\n"
        f"{extracted_text}\n\n"
    )

    response = openai.ChatCompletion.create(
        model="gpt-3.5-turbo",
        messages=[
            {"role": "system", "content": "You are a helpful assistant."},
            {"role": "user", "content": prompt}
        ]
    )

    response_text = response['choices'][0]['message']['content'].strip()
    try:
        data = json.loads(response_text)
        print("Parsed JSON Data:", json.dumps(data, indent=4))
        # return validate_and_correct_items(data)
        return data
    except json.JSONDecodeError as e:
        print("Error decoding JSON:", e)
        raise

def validate_and_correct_items(data):
    """Validate and correct extracted data fields."""
    corrected_items = []
    for item in data.get("items", []):
        try:
            # Attempt to extract and convert the required fields
            description = item.get("Description", "")
            quantity = int(item.get("Quantity", 0))
            rate = float(item.get("Rate", 0))
            sales_amount = float(item.get("Sales Amount", 0))

            # Validate and correct sales amount if necessary
            expected_sales_amount = round(quantity * rate, 2)
            if abs(expected_sales_amount - sales_amount) > 0.01:  # Allow slight rounding errors
                corrected_quantity = max(1, round(sales_amount / rate))  # Avoid zero quantity
                corrected_sales_amount = round(corrected_quantity * rate, 2)
                item["Quantity"] = corrected_quantity
                item["Sales Amount"] = corrected_sales_amount
                #print(f"Corrected {description}: Quantity={corrected_quantity}, Sales Amount={corrected_sales_amount}")
        except (ValueError, TypeError) as e:
            # Log the error and continue processing the next item
            print(f"Error processing item: {item}. Error: {e}")
        
        # Append the item whether corrected or not
        corrected_items.append(item)
    
    # Ensure the corrected items are updated in the data
    data["items"] = corrected_items
    return data

def extract_numeric_quantity(quantity):
    try:
        if isinstance(quantity, (int, float)):  # If quantity is already a number, use it directly
            return int(quantity)
        elif isinstance(quantity, str):  # If it's a string, extract numeric part
            match = re.search(r'\d+', quantity)
            #print(match)
            return int(match.group()) if match else 0  # Return the number if found, otherwise 0
    except (ValueError, TypeError):
        return 0  

def adjust_formula(formula, source_row, target_row):
    if formula and isinstance(formula, str) and formula.startswith('='):
        adjusted_formula = re.sub(r'(\d+)', lambda match: str(int(match.group(0)) + (target_row - source_row)), formula)
        return adjusted_formula
    return formula

def copy_formulas_down(sheet, start_row, end_row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        base_formula = sheet.cell(row=start_row, column=col).value  # Get the formula from the template row
        for row in range(start_row + 1, end_row + 1):
            adjusted_formula = adjust_formula(base_formula, start_row, row)
            sheet.cell(row=row, column=col, value=adjusted_formula)  # Copy the adjusted formula down

def safe_float_conversion(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0
    
def copy_package_weight_formulas(sheet, start_row, end_row, package_weight_col):
    for row in range(start_row, end_row + 1):
        # Set the formula to keep $C$3 fixed
        formula = f"=D{row}*$C$3"
        sheet.cell(row=row, column=package_weight_col).value = formula



def copy_purchase_formula(sheet, start_row, end_row, col_num):
    column_letter = get_column_letter(col_num)
    purchase_formula = "=I{row}*${col_letter}$8"
    for row in range(start_row, end_row + 1):
        formula = purchase_formula.format(row=row, col_letter=column_letter)
        sheet.cell(row=row, column=col_num).value = formula 

def copy_sales_formula(sheet, start_row, end_row, col_num):
    left_column_letter = get_column_letter(col_num - 1)
    sales_formula = "=J{row}*${col_left}$8*(1+${col_left}$7)*(1+$E$10)"
    for row in range(start_row, end_row + 1):
        formula = sales_formula.format(row=row, col_left=left_column_letter)
        sheet.cell(row=row, column=col_num).value = formula  

def copy_mrp_formula(sheet, start_row, end_row, col_num):
    left_column_letter = get_column_letter(col_num - 1)
    mrp_formula = "=INT({col_left}{row}*1.1)+$C$5/100"
    for row in range(start_row, end_row + 1):
        formula = mrp_formula.format(row=row, col_left=left_column_letter)
        sheet.cell(row=row, column=col_num).value = formula 

def load_internal_item_names(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        first_sheet = workbook.worksheets[0]
        internal_item_names = {
            first_sheet[f"A{row}"].value: row for row in range(1, first_sheet.max_row + 1) if first_sheet[f"A{row}"].value
        }
        # for item_name, row in internal_item_names.items():
        #     print(f"Loaded internal item name: '{item_name}' at row {row}")
        return internal_item_names
    except Exception as e:
        print(f"Error loading item names: {e}")
        return {}
    
def load_internal_item_names_list(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        first_sheet = workbook.worksheets[0]
        internal_item_names = [
            first_sheet[f"A{row}"].value for row in range(1, first_sheet.max_row + 1) 
            if first_sheet[f"A{row}"].value
        ]
        return internal_item_names  # Return a list instead of a dictionary
    except Exception as e:
        print(f"Error loading item names: {e}")
        return []  # Return an empty list on error


def update_ann_rate_sheet(ann_rate_sheet, invoice_date, data, item_name_to_row):
    # Finding the next available column to insert new data
    next_available_col = find_next_available_column(ann_rate_sheet)
    print(f"Next available column is {next_available_col} (Column {get_column_letter(next_available_col)})")

    # Write the invoice date in row 2 of this column
    ann_rate_sheet.cell(row=2, column=next_available_col).value = invoice_date
    print(f"Set invoice date '{invoice_date}' at row 2, column {get_column_letter(next_available_col)}")

    # Iterate over each item in the provided data
    for item in data.get("items", []):
        internal_item_name = item.get("internal_item_name", "").strip()  # Normalize the item name
        print(f"Processing item '{internal_item_name}'")

        if internal_item_name in item_name_to_row:
            row_index = max(item_name_to_row[internal_item_name], 3) + 2  # Ensure row starts from at least row 3
            rate = safe_float_conversion(item.get("Rate", 0))  # Safely convert the rate to float
            # Set the rate in the correct row and column
            ann_rate_sheet.cell(row=row_index, column=next_available_col).value = rate
            print(f"Updated row {row_index}, column {get_column_letter(next_available_col)} with rate {rate}")
        else:
            print(f"Item '{internal_item_name}' not found in internal item names.")


def find_next_available_column(sheet):
    col_index = 2  # Starting at column B
    while sheet.cell(row=2, column=col_index).value is not None:
        col_index += 1
    return col_index

def export_to_sigma(parsed_data_list, output_excel_path):
    try:
        workbook = load_workbook(output_excel_path)
        sheet = workbook.active  # Assuming the first sheet is the main sheet where data needs to be added
    except FileNotFoundError:
        print(f"Error: {output_excel_path} not found. Ensure the workbook exists before running this function.")
        return

    # Find the first empty row
    max_row = sheet.max_row
    first_empty_row = max_row + 1 if sheet.cell(row=max_row, column=1).value else max_row

    # Populate data from the parsed_data_list
    for data in parsed_data_list:
        invoice_date = data.get("Date", "Unknown Date")
        invoice_number = data.get("Invoice Number", "Unknown Number")
        supplier_name = data.get("Supplier Name", "Unknown Supplier")

        # Process each item in the invoice
        for index, item in enumerate(data.get("items", []), start=first_empty_row + 1):
            product_name = item.get("internal_item_name", "")
            quantity = item.get("Quantity", 0)
            batch_number = item.get("batchNumber", "00")  # Default to '00' if not found

            # Write data to the corresponding columns in the sheet
            sheet.cell(row=index, column=1, value=invoice_date)
            sheet.cell(row=index, column=2, value=invoice_number)
            sheet.cell(row=index, column=3, value=supplier_name)
            sheet.cell(row=index, column=4, value=batch_number)
            sheet.cell(row=index, column=5, value=product_name)
            sheet.cell(row=index, column=6, value=quantity)

            first_empty_row += 1  # Increment for the next item

    # Save the changes to the workbook
    workbook.save(filename=output_excel_path)
    print(f"Data successfully saved to {output_excel_path}")

def save_to_excel(parsed_data_list, output_excel_path):
    print("writing to excel")
    print(parsed_data_list)
    # Load the existing workbook
    # Load the existing workbook
    try:
        workbook = load_workbook(output_excel_path)
    except FileNotFoundError:
        print(f"Error: The file {output_excel_path} was not found.")
        return

    # Check if the "Template Sheet" exists
    if "Template Sheet" not in workbook.sheetnames:
        print("Template Sheet not found in the workbook.")
        return
    
    if "ANN RATE" not in workbook.sheetnames:
        print("ANN RATE sheet not found in the workbook.")
        return

    template_sheet = workbook["Template Sheet"]
    ann_rate_sheet = workbook["ANN RATE"]

    for data in parsed_data_list:
        supplier_name = data.get("Supplier Name", "Unknown Supplier")
        invoice_date = data.get("Date", "Unknown Date")  # Extract date from invoice number
        sheet_name = f"{supplier_name} {invoice_date}"

        # Ensure the sheet name does not exceed Excel's limit of 31 characters
        sheet_name = sheet_name[:31]

        # Remove existing sheet if it exists to prevent duplicates
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]  # Delete the existing sheet

        # Copy the template sheet to create a new sheet with the desired name
        new_sheet = workbook.copy_worksheet(template_sheet)
        sheet_name = re.sub(r'[\/:*?"<>|]', '_', sheet_name)
        new_sheet.title = sheet_name[:31]

        # Set up fields for each import
        total_weight = 0  # Initialize total weight

        # Calculate total weight for C1 and set other free entry fields
        for item in data.get("items", []):
            rate = safe_float_conversion(item.get("Rate", 0))  # Use safe conversion for Rate
            quantity = extract_numeric_quantity(item.get("Quantity", "0"))
            weight_per_quantity = safe_float_conversion(item.get("weight_per_quantity", 0))
            total_weight += weight_per_quantity * quantity

        # Set values in the Excel sheet as per your format
        new_sheet['C1'] = total_weight  # C1: sum of (each item weight x qty)
        total_transport_cost = safe_float_conversion(data.get("transportCost", 0))  # Fetch the total transportation cost
        new_sheet['C2'] = total_transport_cost  # C2: Total Transportation Cost
        batch_number = data.get("batchNumber", "00")  # Fetch batch number from JSON, defaulting to '00'
        new_sheet['C5'] = batch_number  # C5: Free entry (numbers format limited to 2 digits)

        # Populate item details starting from row 10 without headers
        row = 10
        for index, item in enumerate(data.get("items", []), start=1):
            serial_number = index
            internal_item_name = item.get("internal_item_name", "")  # Use the internal item name from input
            brand_name = item.get("brand_name", "")  # Use the brand name from input
            weight_per_quantity = safe_float_conversion(item.get("weight_per_quantity", 0))  # Use package weight from input
            # weight = safe_float_conversion(item.get("Rate", 0)) * extract_numeric_quantity(item.get("Quantity", "0"))  # D10: Weight = rate * qty
            gst = item.get("GST", "")
            sales_amount = safe_float_conversion(item.get("Sales Amount", ""))
            _quantity = extract_numeric_quantity(item.get("Quantity", "0"))

            # Write data starting from row 10 and in appropriate columns
            new_sheet.cell(row=row, column=1, value=serial_number)  # Column A
            new_sheet.cell(row=row, column=2, value=internal_item_name)  # Column B
            new_sheet.cell(row=row, column=3, value=brand_name)  # Column C
            new_sheet.cell(row=row, column=4, value=weight_per_quantity * _quantity)  # Column D
            print(f"Weight per quantity: {weight_per_quantity}, Quantity: {_quantity}") #remove this
            new_sheet.cell(row=row, column=5, value=f"{gst}%")  # Column E (Formatted with %)
            new_sheet.cell(row=row, column=6, value=sales_amount)  # Column F
            #new_sheet.cell(row=row, column=7, value=package_weight)  # Column G (Package Weight)

            # Continue to next row
            row += 1
        item_name_to_row = load_internal_item_names("output/3 - Internal Item Name List.xlsx")
        update_ann_rate_sheet(ann_rate_sheet, invoice_date, data, item_name_to_row)

        export_to_sigma(parsed_data_list, "output/5 - Product List in Sigma.xlsx")

        copy_package_weight_formulas(new_sheet, start_row=10, end_row=row - 1, package_weight_col=7)  # Special handling for package weight
        copy_formulas_down(new_sheet, start_row=10, end_row=row - 1, start_col=8, end_col=10)  # Copy other formulas
        
        columns = [
        (12, 13, 14),  # L, M, N
        (15, 16, 17),  # O, P, Q
        (18, 19, 20),  # R, S, T
        (21, 22, 23),  # U, V, W
        (24, 25, 26),  # X, Y, Z
        (27, 28, 29),  # AA, AB, AC
        (30, 31, 32)   # AD, AE, AF
        ]

        for purchase_col, sales_col, mrp_col in columns:
            # Call the functions with the specified column numbers
            copy_purchase_formula(new_sheet, 10, row - 1, purchase_col)
            copy_sales_formula(new_sheet, 10, row - 1, sales_col)
            copy_mrp_formula(new_sheet, 10, row - 1, mrp_col)


    # Save the workbook back to the same file to update in place
    try:
        workbook.save(output_excel_path)
        print(f"Data successfully updated in {output_excel_path}")
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return

    if os.path.exists(output_excel_path):
        print(f"File exists after saving: {output_excel_path}")
        open_excel_file(os.path.abspath('output/4 - Master Excel File.xlsx'))  # Attempt to open the file
    else:
        print(f"File does not exist after saving: {output_excel_path}")

@app.route('/', methods=['GET', 'POST'])
def index():
    internal_item_names = load_internal_item_names_list("output/3 - Internal Item Name List.xlsx")  # Load names
    processed_files = []  # Initialize as empty for GET requests
    all_invoices = []  # This will hold all invoices for merging later

    # Debugging: Log internal_item_names
    print("Loaded Internal Item Names:", internal_item_names)

    if request.method == 'POST':
        files = request.files.getlist('files[]')
        for file in files:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                file.save(filepath)

                # Handle PDFs by splitting them into pages
                if filepath.lower().endswith('.pdf'):
                    page_files = split_pdf_into_pages(filepath)
                else:
                    page_files = [filepath]  # Treat non-PDF files as single files

                # Process each page or file separately
                for page_file in page_files:
                    try:
                        # Perform OCR and parse the invoice
                        filepath = convert_to_jpg(page_file)  # Use existing convert_to_jpg
                        extracted_text = perform_ocr(filepath)
                        parsed_data = parse_invoice_with_genai(extracted_text)

                        # Append the parsed data to all_invoices for later merging
                        all_invoices.append(parsed_data)

                    except ValueError as e:
                        flash(str(e), 'danger')
                        return redirect(url_for('index'))

        # Merge invoices by supplier before rendering or saving
        merged_invoices = merge_invoice_data(all_invoices)

        # Save the merged invoices for review
        for invoice in merged_invoices:
            processed_file = {
                'filename': invoice.get("Invoice Number", "Unknown Invoice"),
                'data': invoice
            }
            processed_files.append(processed_file)

        # Render the review page with processed files and internal item names
        return render_template('review.html', processed_files=processed_files, internal_item_names=internal_item_names)

    # Render the index page with internal item names
    return render_template('index.html', internal_item_names=internal_item_names)  # Pass names on GET



@app.route('/accept', methods=['POST'])
def accept():
    parsed_data_list = request.form['parsed_data']
    try:
        parsed_data_json = json.loads(parsed_data_list)
        processed_files = parsed_data_json.get('processedFiles', [])
    except json.JSONDecodeError as e:
        flash(f"Error decoding JSON data: {e}", 'danger')
        return redirect(url_for('index'))
    
    invoice_data = [file['data'] for file in processed_files]

    # Save parsed data to Excel
    output_excel_path = os.path.join(app.config['OUTPUT_FOLDER'], '4 - Master Excel File.xlsx')
    save_to_excel(invoice_data, output_excel_path)
    
    # Instead of sending a file, redirect to a confirmation page or back to the form with a success message
    flash("Data has been successfully saved to Excel.", 'success')
    return redirect(url_for('index'))  # Assuming 'index' is the route

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'pdf'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

if __name__ == "__main__":
    app.run(debug=True)